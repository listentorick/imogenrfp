import os
import time
import logging
from typing import Optional
from pathlib import Path
from sqlalchemy.orm import Session
from database import get_db
from models import Document
from queue_service import queue_service
from websocket_manager import websocket_manager
from chroma_service import chroma_service
from question_extraction_service import question_extraction_service

# Document processing libraries
import PyPDF2
from docx import Document as DocxDocument
import magic
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from langchain_text_splitters import RecursiveCharacterTextSplitter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    def __init__(self):
        # We'll use the chroma_service instead of direct client
        self.chroma_service = chroma_service
        
        # Initialize LangChain text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""]  # Prioritize paragraph, sentence, then word breaks
        )
    
    def extract_text_from_file(self, file_path: str) -> str:
        """Extract text content from various file types"""
        try:
            # Detect file type
            mime_type = magic.from_file(file_path, mime=True)
            
            if mime_type == 'application/pdf':
                return self._extract_from_pdf(file_path)
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 
                              'application/msword']:
                return self._extract_from_docx(file_path)
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'application/vnd.ms-excel']:
                excel_data = self._extract_from_excel(file_path)
                return excel_data.get('text_content', '')
            elif mime_type.startswith('text/'):
                return self._extract_from_text(file_path)
            else:
                logger.warning(f"Unsupported file type: {mime_type}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            return ""
    
    def extract_excel_data(self, file_path: str) -> dict:
        """Extract structured Excel data for question extraction"""
        try:
            mime_type = magic.from_file(file_path, mime=True)
            if mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           'application/vnd.ms-excel']:
                return self._extract_from_excel(file_path)
            return {}
        except Exception as e:
            logger.error(f"Error extracting Excel data from {file_path}: {e}")
            return {}
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {e}")
        return text
    
    def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = DocxDocument(file_path)
            return "\n".join([paragraph.text for paragraph in doc.paragraphs])
        except Exception as e:
            logger.error(f"Error reading DOCX {file_path}: {e}")
            return ""
    
    def _extract_from_text(self, file_path: str) -> str:
        """Extract text from plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error reading text file {file_path}: {e}")
            return ""
    
    def _extract_from_excel(self, file_path: str) -> dict:
        """Extract structured data from Excel file"""
        try:
            # Load workbook to get sheet names and preserve cell references
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active  # For now, assume single sheet
            
            # Extract cell data with coordinates
            cells_data = []
            text_content = ""
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                        cell_value = str(cell.value).strip()
                        cells_data.append({
                            'cell_reference': cell_ref,
                            'value': cell_value,
                            'row': cell.row,
                            'column': cell.column
                        })
                        text_content += f"Cell {cell_ref}: {cell_value}\n"
            
            logger.info(f"Extracted {len(cells_data)} non-empty cells from Excel file")
            
            return {
                'text_content': text_content,
                'cells_data': cells_data,
                'sheet_name': sheet.title,
                'total_cells': len(cells_data)
            }
        except Exception as e:
            logger.error(f"Error reading Excel {file_path}: {e}")
            return {'text_content': '', 'cells_data': [], 'sheet_name': '', 'total_cells': 0}
    
    def store_in_vector_db(self, project_id: str, document_id: str, text: str, metadata: dict):
        """Store document text in ChromaDB using chroma_service"""
        try:
            # Split text into chunks using LangChain's RecursiveCharacterTextSplitter
            chunks = self.text_splitter.split_text(text)
            
            # Use chroma_service to store the document
            success = self.chroma_service.add_document_to_project(
                project_id=project_id,
                document_id=document_id,
                text_chunks=chunks,
                metadata=metadata
            )
            
            if success:
                logger.info(f"Stored document {document_id} in vector DB with {len(chunks)} chunks")
            else:
                logger.error(f"Failed to store document {document_id} in vector DB")
            
            return success
        except Exception as e:
            logger.error(f"Error storing document {document_id} in vector DB: {e}")
            return False
    
    def update_document_status(self, document_id: str, status: str, error_message: Optional[str] = None, tenant_id: str = None):
        """Update document status in database and notify via WebSocket"""
        db = next(get_db())
        try:
            document = db.query(Document).filter(Document.id == document_id).first()
            if document:
                document.status = status
                if error_message:
                    document.processing_error = error_message
                elif status == 'processed':
                    # Clear error message when processing succeeds
                    document.processing_error = None
                db.commit()
                logger.info(f"Updated document {document_id} status to {status}")
                
                # Publish status update via WebSocket
                if tenant_id:
                    websocket_manager.publish_document_status_update(
                        tenant_id=tenant_id,
                        document_id=document_id,
                        status=status,
                        error_message=error_message
                    )
        except Exception as e:
            logger.error(f"Error updating document status: {e}")
            db.rollback()
        finally:
            db.close()
    
    def process_document(self, job_data: dict):
        """Process a single document"""
        document_id = job_data['document_id']
        file_path = job_data['file_path']
        tenant_id = job_data['tenant_id']
        project_id = job_data['project_id']
        deal_id = job_data.get('deal_id')  # Deal ID is optional
        
        logger.info(f"Processing document {document_id} at {file_path}")
        
        try:
            # Update status to processing
            self.update_document_status(document_id, 'processing', tenant_id=tenant_id)
            
            # Get document info from database to access original filename
            db = next(get_db())
            try:
                document = db.query(Document).filter(Document.id == document_id).first()
                original_filename = document.original_filename if document else os.path.basename(file_path)
            finally:
                db.close()
            
            # Extract text from file
            text = self.extract_text_from_file(file_path)
            
            if not text.strip():
                raise Exception("No text content extracted from file")
            
            # Only store project documents in ChromaDB, not deal documents
            if project_id and not deal_id:
                # This is a project document - store in ChromaDB
                logger.info(f"Storing project document {document_id} ({original_filename}) in ChromaDB")
                
                # Ensure project collection exists
                self.chroma_service.create_project_collection(project_id, f"Project {project_id}")
                
                # Store in project-specific ChromaDB collection
                success = self.store_in_vector_db(
                    project_id=project_id,
                    document_id=document_id,
                    text=text,
                    metadata={
                        'tenant_id': tenant_id,
                        'project_id': project_id,
                        'document_id': document_id,
                        'file_path': file_path,
                        'filename': original_filename
                    }
                )
                
                if not success:
                    raise Exception("Failed to store in vector database")
            else:
                # This is a deal document - skip ChromaDB storage
                logger.info(f"Skipping ChromaDB storage for deal document {document_id}")
                success = True  # No storage needed for deal documents
            
            # Extract questions if document is associated with a deal
            if deal_id:
                logger.info(f"Extracting questions from document {document_id} for deal {deal_id}")
                
                # For Excel documents, get structured data for better question extraction
                excel_data = None
                if file_path:
                    try:
                        mime_type = magic.from_file(file_path, mime=True)
                        if mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                       'application/vnd.ms-excel']:
                            excel_data = self.extract_excel_data(file_path)
                            logger.info(f"Extracted Excel data with {excel_data.get('total_cells', 0)} cells")
                    except Exception as e:
                        logger.warning(f"Could not extract Excel data: {e}")
                
                question_success = question_extraction_service.process_document_for_questions(
                    document_id, text, excel_data
                )
                if question_success:
                    logger.info(f"Successfully extracted questions from document {document_id}")
                else:
                    logger.warning(f"Failed to extract questions from document {document_id}")
                    # Don't fail the entire process if question extraction fails
            
            self.update_document_status(document_id, 'processed', tenant_id=tenant_id)
            logger.info(f"Successfully processed document {document_id}")
                
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error processing document {document_id}: {error_msg}")
            self.update_document_status(document_id, 'error', error_msg, tenant_id=tenant_id)

def run_document_processor():
    """Main worker loop"""
    processor = DocumentProcessor()
    logger.info("Document processor started")
    
    while True:
        try:
            # Get next job from queue
            job_data = queue_service.dequeue_document_processing()
            
            if job_data:
                processor.process_document(job_data)
            else:
                # No jobs available, wait a bit
                time.sleep(1)
                
        except KeyboardInterrupt:
            logger.info("Document processor stopped")
            break
        except Exception as e:
            logger.error(f"Error in document processor: {e}")
            time.sleep(5)  # Wait before retrying

if __name__ == "__main__":
    run_document_processor()