import logging
import os
import time
import shutil
from datetime import datetime
from typing import Dict, Any
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from pathlib import Path

from database import get_db
from models import Export, Document, Question, Deal
from export_service import export_service

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExportWorker:
    def __init__(self):
        self.export_dir = os.getenv('EXPORT_DIR', '/app/exports')
        # Ensure export directory exists
        Path(self.export_dir).mkdir(parents=True, exist_ok=True)
    
    def process_export_job(self, job: Dict[str, Any]) -> bool:
        """Process a single export job"""
        export_id = job.get('export_id')
        tenant_id = job.get('tenant_id')
        deal_id = job.get('deal_id')
        document_id = job.get('document_id')
        
        logger.info(f"Processing export job {export_id}")
        
        db = next(get_db())
        try:
            # Get export record
            export_record = db.query(Export).filter(Export.id == export_id).first()
            if not export_record:
                logger.error(f"Export record {export_id} not found")
                return False
            
            # Update status to processing
            export_record.status = 'processing'
            db.commit()
            
            # Get source document
            document = db.query(Document).filter(Document.id == document_id).first()
            if not document:
                logger.error(f"Source document {document_id} not found")
                self._mark_export_failed(db, export_record, "Source document not found")
                return False
            
            # Get questions for this deal
            questions = db.query(Question).filter(
                Question.deal_id == deal_id,
                Question.tenant_id == tenant_id
            ).order_by(Question.question_order).all()
            
            logger.info(f"Found {len(questions)} questions for export")
            
            # Process based on document type
            if document.original_filename.lower().endswith(('.xlsx', '.xls')):
                success = self._process_excel_export(db, export_record, document, questions)
            else:
                # For non-Excel documents, we'll create a simple text file for now
                success = self._process_text_export(db, export_record, document, questions)
            
            if success:
                export_record.status = 'completed'
                export_record.completed_at = datetime.utcnow()
                db.commit()
                logger.info(f"Export {export_id} completed successfully")
            else:
                self._mark_export_failed(db, export_record, "Processing failed")
            
            return success
            
        except Exception as e:
            logger.error(f"Error processing export job {export_id}: {e}")
            self._mark_export_failed(db, export_record, str(e))
            return False
        finally:
            db.close()
    
    def _process_excel_export(self, db: Session, export_record: Export, document: Document, questions: list) -> bool:
        """Process Excel document export with cell mapping"""
        try:
            # Copy source file to export directory
            source_path = document.file_path
            export_filename = f"export_{export_record.id}_{document.original_filename}"
            export_path = os.path.join(self.export_dir, export_filename)
            
            # Check if source file exists
            if not os.path.exists(source_path):
                logger.error(f"Source file does not exist: {source_path}")
                return False
            
            logger.info(f"Copying {source_path} to {export_path}")
            shutil.copy2(source_path, export_path)
            
            # Load the Excel workbook
            workbook = load_workbook(export_path)
            
            # Track answered questions
            answered_count = 0
            
            # Process each question
            for question in questions:
                if question.answer_text and question.answer_cell_reference:
                    try:
                        # Determine which sheet to use
                        sheet_name = question.sheet_name if question.sheet_name else workbook.active.title
                        
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            # Write answer to the specified cell
                            sheet[question.answer_cell_reference] = question.answer_text
                            answered_count += 1
                            
                            logger.info(f"Wrote answer to cell {question.answer_cell_reference} in sheet {sheet_name}")
                        else:
                            logger.warning(f"Sheet {sheet_name} not found in workbook")
                            
                    except Exception as e:
                        logger.error(f"Error writing answer for question {question.id}: {e}")
                elif question.answer_text:
                    logger.warning(f"Question has answer but no cell reference: {question.question_text[:50]}...")
            
            # Save the modified workbook
            workbook.save(export_path)
            workbook.close()
            
            # Update export record
            export_record.file_path = export_path
            export_record.export_filename = export_filename
            export_record.answered_count = answered_count
            
            logger.info(f"Excel export completed: {answered_count} answers written to {export_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error in Excel export processing: {e}")
            return False
    
    def _process_text_export(self, db: Session, export_record: Export, document: Document, questions: list) -> bool:
        """Process non-Excel document export as text file"""
        try:
            export_filename = f"export_{export_record.id}_{document.original_filename}.txt"
            export_path = os.path.join(self.export_dir, export_filename)
            
            answered_count = 0
            
            with open(export_path, 'w', encoding='utf-8') as f:
                f.write(f"Export Report for {document.original_filename}\n")
                f.write(f"Generated: {datetime.utcnow().isoformat()}\n")
                f.write("=" * 50 + "\n\n")
                
                for i, question in enumerate(questions, 1):
                    f.write(f"Question {i}: {question.question_text}\n")
                    
                    if question.answer_text:
                        f.write(f"Answer: {question.answer_text}\n")
                        answered_count += 1
                    else:
                        f.write("Answer: [Not answered]\n")
                    
                    if question.answer_relevance_score:
                        f.write(f"Relevance Score: {question.answer_relevance_score}%\n")
                    
                    f.write("\n" + "-" * 30 + "\n\n")
            
            # Update export record
            export_record.file_path = export_path
            export_record.export_filename = export_filename
            export_record.answered_count = answered_count
            
            logger.info(f"Text export completed: {answered_count} answers written to {export_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error in text export processing: {e}")
            return False
    
    def _mark_export_failed(self, db: Session, export_record: Export, error_message: str):
        """Mark export as failed with error message"""
        try:
            export_record.status = 'failed'
            export_record.error_message = error_message
            export_record.completed_at = datetime.utcnow()
            db.commit()
        except Exception as e:
            logger.error(f"Error marking export as failed: {e}")
    
    def run(self):
        """Main worker loop"""
        logger.info("Export worker starting...")
        
        while True:
            try:
                # Get next job from queue
                job = export_service.dequeue_export_job()
                
                if job:
                    self.process_export_job(job)
                else:
                    # No job available, sleep briefly
                    time.sleep(1)
                    
            except KeyboardInterrupt:
                logger.info("Export worker shutting down...")
                break
            except Exception as e:
                logger.error(f"Unexpected error in worker loop: {e}")
                time.sleep(5)  # Wait before retrying

if __name__ == '__main__':
    worker = ExportWorker()
    worker.run()